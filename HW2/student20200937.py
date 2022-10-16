#!/usr/bin/python3

from openpyxl import load_workbook

wb = load_workbook(filename = 'student.xlsx')
ws = wb['Sheet1']

num = ws.max_row - 1

for row in range(2, num + 2):
	total = 0
	total += ws.cell(column = 3, row = row).value * 0.3
	total += ws.cell(column = 4, row = row).value * 0.35
	total += ws.cell(column = 5, row = row).value * 0.34
	total += 1
	ws.cell(column = 7, row = row, value = round(total, 2))

score = []
for row in range(2, num + 2):
	score.append(ws.cell(column = 7, row = row).value)
	
score.sort(reverse = True)

cnt = 0
rank = 0

for i in range(num):
	cnt = score.count(score[i])
	rank += cnt
	for row in range(2, num + 2):
		if ws.cell(column = 7, row = row).value == score[i]:
			ws.cell(column = 9, row = row).value = rank
	i += cnt
	
a=0
b=0
c=0
aMax=0
bMax=0
for row in range(2, num + 2):
	rank = ws.cell(column = 9, row = row).value/num*100
	if rank <= 30:
		ws.cell(column = 8, row = row).value = "A0"
		a += 1
		if aMax<rank:
			aMax = rank
	elif rank <= 70:
		ws.cell(column = 8, row = row).value = "B0"
		b += 1
		if bMax<rank:
			bMax = rank
	else:
		ws.cell(column = 8, row = row).value = "C0"
		c += 1
		
ap = a/num*50
bp = b/num*50 + aMax
cp = c/num*50 + bMax

for row in range(2, num + 2):
	rank = ws.cell(column = 9, row = row).value/num*100
	if rank <= ap:
		ws.cell(column = 8, row = row).value = "A+"
	elif rank > aMax and rank <= bp:
		ws.cell(column = 8, row = row).value = "B+"
	elif rank > bMax and rank <= cp:
		ws.cell(column = 8, row = row).value = "C+"
	
ws.delete_cols(9)

wb.save("student.xlsx")
